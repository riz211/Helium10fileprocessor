from typing import Optional, Tuple
import pandas as pd
from openpyxl.styles import PatternFill, Font
from io import BytesIO

def read_excel_file(file: BytesIO) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """Read an Excel file and return DataFrame or error message."""
    try:
        excel_file = pd.ExcelFile(file)
        dfs = []

        for sheet_name in excel_file.sheet_names:
            # Read all columns first
            sheet_data = pd.read_excel(file, sheet_name=sheet_name)

            # Define multiple possible column mappings
            column_mappings = [
                {
                    "Product Details": "TITLE",
                    "Brand": "BRAND",
                    "Product ID": "SKU",
                    "UPC Code": "UPC/ISBN",
                    "Price ": "COST_PRICE"  # Note the space after "Price"
                },
                {
                    "TITLE": "TITLE",
                    "Brand": "BRAND",
                    "SKU": "SKU",
                    "UPC/ISBN": "UPC/ISBN",
                    "COST_PRICE": "COST_PRICE"
                }
            ]

            # Try each mapping
            for mapping in column_mappings:
                # Check if all required columns from this mapping exist
                if all(col in sheet_data.columns for col in mapping.keys()):
                    # Select and rename columns
                    selected_data = sheet_data[list(mapping.keys())].copy()
                    selected_data.rename(columns=mapping, inplace=True)
                    dfs.append(selected_data)
                    break  # Found a working mapping, stop checking others
            else:
                # None of the mappings worked, construct detailed error message
                available_cols = set(sheet_data.columns)
                missing_cols = []
                for mapping in column_mappings:
                    missing = [col for col in mapping.keys() if col not in available_cols]
                    if missing:
                        missing_cols.extend(missing)

                missing_cols = list(set(missing_cols))  # Remove duplicates
                return None, f"Required columns not found in sheet '{sheet_name}'. Missing one of these variations: {', '.join(missing_cols)}"

        if not dfs:
            return None, "No valid data found in any sheet"

        return pd.concat(dfs, ignore_index=True), None

    except Exception as e:
        return None, f"Error reading file: {str(e)}"

def calculate_shipping_cost(weight: float, legend: pd.DataFrame) -> Optional[float]:
    """Calculate shipping cost based on weight and shipping legend."""
    if pd.isnull(weight):
        return None

    try:
        for _, row in legend.iterrows():
            if row["Weight Range Min (lb)"] <= weight <= row["Weight Range Max (lb)"]:
                # Keep original value, format to 1 decimal place
                return float(row["SHIPPING COST"])
        return None
    except Exception:
        return None

def create_excel_export(
    df: pd.DataFrame,
    shipping_legend: Optional[pd.DataFrame] = None
) -> BytesIO:
    """Create formatted Excel export with optional shipping legend."""
    buffer = BytesIO()

    try:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Write main data
            df.to_excel(writer, index=False, sheet_name="Consolidated Data")

            # Add shipping legend if provided
            if shipping_legend is not None:
                shipping_legend.to_excel(writer, index=False, sheet_name="ShippingLegend")

            # Format worksheet
            worksheet = writer.sheets["Consolidated Data"]

            # Format headers
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            # Initialize columns dictionary
            column_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}

            # Define fill colors
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

            # Step 12.4: Format numeric columns to 2 decimal places
            numeric_columns = [
                "COST_PRICE",
                "HANDLING COST",
                "ITEM WEIGHT (pounds)",
                "SHIPPING COST",
                "RETAIL PRICE",
                "MIN PRICE",
                "MAX PRICE",
            ]
            for col in numeric_columns:
                col_index = column_indices.get(col)
                if col_index:
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_index)
                        cell.number_format = '0.00'


            # Set number format for shipping cost column
            shipping_cost_col = column_indices.get("SHIPPING COST")
            if shipping_cost_col:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=shipping_cost_col)
                    cell.number_format = '0.0'  # Format to show 1 decimal place

            # Set number format for RETAIL PRICE column to show 2 decimal places
            retail_price_col = column_indices.get("RETAIL PRICE")
            if retail_price_col:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=retail_price_col)
                    cell.number_format = '0.00'  # Format to show 2 decimal places

            # Get column indices for formulas
            weight_col = column_indices.get("ITEM WEIGHT (pounds)")
            shipping_cost_col = column_indices.get("SHIPPING COST")
            cost_price_col = column_indices.get("COST_PRICE")
            handling_cost_col = column_indices.get("HANDLING COST")
            retail_price_col = column_indices.get("RETAIL PRICE")

            # Iterate over rows and apply formatting and formulas
            for row_idx in range(2, worksheet.max_row + 1):
                # Get weight and retail price values
                weight_cell = worksheet.cell(row=row_idx, column=weight_col)
                retail_price_cell = worksheet.cell(row=row_idx, column=retail_price_col) if retail_price_col else None

                # Check conditions for highlighting
                is_missing_weight = weight_cell.value is None or weight_cell.value == ""
                has_low_retail_price = (retail_price_cell and retail_price_cell.value is not None
                                      and isinstance(retail_price_cell.value, (int, float))
                                      and float(retail_price_cell.value) < 10)

                # Apply highlighting
                if is_missing_weight:
                    fill = red_fill
                elif has_low_retail_price:
                    fill = orange_fill
                else:
                    fill = None

                if fill:
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = fill

                # Add formulas for rows with missing weights
                if is_missing_weight:
                    # Add shipping cost formula
                    if shipping_cost_col:
                        shipping_formula = f'=IF(ISBLANK({get_column_letter(weight_col)}{row_idx}),"",ROUND(VLOOKUP({get_column_letter(weight_col)}{row_idx},ShippingLegend!A:C,3,TRUE),1))'
                        worksheet.cell(row=row_idx, column=shipping_cost_col).value = shipping_formula

                    # Add retail price formula
                    if retail_price_col and cost_price_col and handling_cost_col and shipping_cost_col:
                        retail_formula = (
                            f'=IF(AND('
                            f'{get_column_letter(cost_price_col)}{row_idx}<>"",'
                            f'{get_column_letter(handling_cost_col)}{row_idx}<>"",'
                            f'{get_column_letter(shipping_cost_col)}{row_idx}<>""),'
                            f'ROUND(('
                            f'{get_column_letter(cost_price_col)}{row_idx}+'
                            f'{get_column_letter(handling_cost_col)}{row_idx}+'
                            f'{get_column_letter(shipping_cost_col)}{row_idx})*1.35,2),"")'
                        )
                        worksheet.cell(row=row_idx, column=retail_price_col).value = retail_formula

                        # Add MIN PRICE formula (equal to RETAIL PRICE)
                        min_price_col = column_indices.get("MIN PRICE")
                        if min_price_col:
                            worksheet.cell(row=row_idx, column=min_price_col).value = f"=K{row_idx}"  # Assuming RETAIL PRICE is column K

                        # Add MAX PRICE formula (MIN PRICE * 1.35)
                        max_price_col = column_indices.get("MAX PRICE")
                        if max_price_col:
                            worksheet.cell(row=row_idx, column=max_price_col).value = f"=IF(L{row_idx}<>\"\", ROUND(L{row_idx}*1.35, 2),\"\")"  # Assuming MIN PRICE is column L

        buffer.seek(0)
        return buffer

    except Exception as e:
        raise Exception(f"Error creating Excel export: {str(e)}")

def get_column_letter(col_num):
    """Convert column number to letter (1 = A, 2 = B, etc.)"""
    result = ""
    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result